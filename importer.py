from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from .db import Database, RecurringCharge
from .localization import CATEGORY_TEXT

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for Excel import
    load_workbook = None


EXCEL_TRANSACTION_SHEET_NAMES = {"transactions"}
EXCEL_SETTINGS_SHEET_NAMES = {"settings", "reglages"}
EXCEL_EXPORT_DATA_SHEET_NAMES = {"export data"}
EXCEL_RECURRING_SHEET_NAMES = {"recurring", "recurring charges", "charges recurrentes"}
EXCEL_CATEGORY_SHEET_NAMES = {"categories"}
EXCEL_TRANSACTION_HEADER_MAP = {
    "date": "datetime_local",
    "datetime": "datetime_local",
    "datetime local": "datetime_local",
    "datetime_local": "datetime_local",
    "type": "type",
    "amount": "amount",
    "montant": "amount",
    "label": "label",
    "libelle": "label",
    "category": "category",
    "categorie": "category",
    "notes": "notes",
    "id": "id",
    "uuid": "id",
    "transaction id": "id",
    "transaction_id": "id",
    "transactionid": "id",
    "recurring id": "recurring_id",
    "recurring_id": "recurring_id",
    "recurringid": "recurring_id",
    "id recurrent": "recurring_id",
    "excluded from averages": "excluded_from_averages",
    "exclude from averages": "excluded_from_averages",
    "avg excluded": "excluded_from_averages",
    "exclu des moyennes": "excluded_from_averages",
}
EXCEL_RECURRING_HEADER_MAP = {
    "id": "id",
    "uuid": "id",
    "label": "label",
    "libelle": "label",
    "amount": "amount",
    "montant": "amount",
    "type": "type",
    "category": "category",
    "categorie": "category",
    "notes": "notes",
    "start date": "start_date",
    "start_date": "start_date",
    "date debut": "start_date",
    "end date": "end_date",
    "end_date": "end_date",
    "date fin": "end_date",
    "status": "status",
    "statut": "status",
    "frequency": "frequency",
    "frequence": "frequency",
    "last applied": "last_applied",
    "last applied date": "last_applied",
    "last_applied": "last_applied",
    "derniere application": "last_applied",
}
EXCEL_SETTINGS_LABEL_MAP = {
    "starting balance": "starting_balance",
    "solde initial": "starting_balance",
    "refunds total": "refunds_total",
    "total remboursements": "refunds_total",
    "plan months total": "plan_months_total",
    "total mois plan": "plan_months_total",
    "months elapsed": "months_elapsed",
    "mois ecoules": "months_elapsed",
    "rent monthly (manual)": "rent_monthly_manual",
    "loyer mensuel (manuel)": "rent_monthly_manual",
    "food house monthly": "food_house_monthly",
    "nourriture maison mensuel": "food_house_monthly",
    "medical monthly": "medical_monthly",
    "medical mensuel": "medical_monthly",
    "school monthly": "school_monthly",
    "ecole mensuel": "school_monthly",
    "household monthly": "household_monthly",
    "maison mensuel": "household_monthly",
    "health monthly": "health_monthly",
    "sante mensuel": "health_monthly",
    "misc monthly": "misc_monthly",
    "divers mensuel": "misc_monthly",
    "extra monthly (12 mo)": "extra_monthly",
    "extra mensuel (12 mois)": "extra_monthly",
    "campus reference total": "campus_reference_total",
    "reference campus": "campus_reference_total",
    "language": "language",
    "langue": "language",
}
EXCEL_INT_SETTINGS = {"plan_months_total", "months_elapsed"}
EXCEL_FLOAT_SETTINGS = {
    "starting_balance",
    "refunds_total",
    "rent_monthly_manual",
    "food_house_monthly",
    "medical_monthly",
    "school_monthly",
    "household_monthly",
    "health_monthly",
    "misc_monthly",
    "extra_monthly",
    "campus_reference_total",
}
TRANSACTION_TYPE_ALIASES = {
    "expense": "expense",
    "depense": "expense",
    "income": "income",
    "revenu": "income",
}
CATEGORY_LOOKUP = {
    normalized: canonical
    for canonical, labels in CATEGORY_TEXT.items()
    for normalized in {
        canonical.lower(),
        *(str(label).strip().lower() for label in labels.values()),
    }
}


@dataclass
class ImportResult:
    added: int
    skipped: int
    recurring_added: int
    recurring_skipped: int
    categories_added: int
    settings_updated: bool
    messages: List[str]
    invalid_transactions: int = 0
    invalid_recurring: int = 0


def import_from_directory(db: Database, directory: Path) -> ImportResult:
    json_path, csv_path, excel_path = _discover_files(directory)
    return import_from_files(
        db,
        json_path=json_path,
        csv_path=csv_path,
        excel_path=excel_path,
    )


def import_from_files(
    db: Database,
    json_path: Optional[Path] = None,
    csv_path: Optional[Path] = None,
    excel_path: Optional[Path] = None,
) -> ImportResult:
    messages: List[str] = []
    settings_updates: Dict[str, object] = {}
    candidates: List[Dict[str, object]] = []
    recurring_candidates: List[RecurringCharge] = []
    category_candidates: List[str] = []
    invalid_transactions = 0
    invalid_recurring = 0

    if json_path:
        (
            settings_updates,
            json_transactions,
            json_recurring,
            json_categories,
            json_messages,
            invalid_json_transactions,
            invalid_json_recurring,
        ) = _load_json(json_path)
        candidates.extend(json_transactions)
        recurring_candidates.extend(json_recurring)
        category_candidates.extend(json_categories)
        messages.extend(json_messages)
        invalid_transactions += invalid_json_transactions
        invalid_recurring += invalid_json_recurring
        messages.append(f"Loaded JSON: {json_path.name}")
    if csv_path:
        csv_transactions, csv_messages, invalid_csv_transactions = _load_csv(csv_path)
        candidates.extend(csv_transactions)
        messages.extend(csv_messages)
        invalid_transactions += invalid_csv_transactions
        messages.append(f"Loaded CSV: {csv_path.name}")
    if excel_path:
        (
            excel_settings_updates,
            excel_transactions,
            excel_recurring,
            excel_categories,
            excel_messages,
            invalid_excel_transactions,
            invalid_excel_recurring,
        ) = _load_excel(excel_path)
        candidates.extend(excel_transactions)
        recurring_candidates.extend(excel_recurring)
        category_candidates.extend(excel_categories)
        messages.extend(excel_messages)
        invalid_transactions += invalid_excel_transactions
        invalid_recurring += invalid_excel_recurring
        for key, value in excel_settings_updates.items():
            settings_updates.setdefault(key, value)
        messages.append(f"Loaded Excel: {excel_path.name}")

    if not candidates and not recurring_candidates and not settings_updates and not category_candidates:
        raise ValueError("No importable data found.")

    existing_transactions = db.list_transactions()
    existing = _existing_fingerprints(existing_transactions)
    existing_ids = {
        str(tx.id).strip() for tx in existing_transactions if str(tx.id).strip()
    }
    new_transactions = []
    skipped = 0
    seen = set(existing)
    seen_ids = set(existing_ids)
    for tx in candidates:
        tx_id = _normalize_optional_text(tx.get("id"))
        if tx_id and tx_id in seen_ids:
            skipped += 1
            continue
        fingerprint = _fingerprint(tx)
        if fingerprint in seen:
            skipped += 1
            continue
        seen.add(fingerprint)
        if tx_id:
            seen_ids.add(tx_id)
        new_transactions.append(tx)

    if new_transactions:
        db.add_transactions_bulk(new_transactions, backup=False, preserve_ids=True)

    category_candidates.extend(
        str(tx.get("category", "")).strip() for tx in new_transactions
    )

    recurring_existing = db.list_recurring_charges()
    existing_recurring_ids = {charge.id for charge in recurring_existing}
    existing_recurring_fingerprints = {
        _recurring_fingerprint(charge) for charge in recurring_existing
    }
    recurring_added = 0
    recurring_skipped = 0
    for charge in recurring_candidates:
        if charge.id in existing_recurring_ids:
            recurring_skipped += 1
            continue
        fingerprint = _recurring_fingerprint(charge)
        if fingerprint in existing_recurring_fingerprints:
            recurring_skipped += 1
            continue
        db.insert_recurring_charge_with_id(charge, backup=False)
        existing_recurring_ids.add(charge.id)
        existing_recurring_fingerprints.add(fingerprint)
        recurring_added += 1
        category_candidates.append(charge.category)

    categories_added = db.ensure_categories(category_candidates, backup=False)

    settings_updated = False
    if settings_updates:
        db.update_settings(settings_updates, backup=False)
        settings_updated = True

    if (
        new_transactions
        or settings_updates
        or recurring_added
        or categories_added
    ):
        db.backup_now()

    return ImportResult(
        added=len(new_transactions),
        skipped=skipped,
        recurring_added=recurring_added,
        recurring_skipped=recurring_skipped,
        categories_added=categories_added,
        settings_updated=settings_updated,
        messages=messages,
        invalid_transactions=invalid_transactions,
        invalid_recurring=invalid_recurring,
    )


def _discover_files(directory: Path) -> Tuple[Optional[Path], Optional[Path], Optional[Path]]:
    if not directory.exists():
        raise FileNotFoundError(str(directory))
    json_path = directory / "budget_history_export.json"
    csv_path = directory / "transactions_export.csv"
    excel_path = directory / "budget_export.xlsx"

    if not json_path.exists():
        json_candidates = sorted(directory.glob("*.json"))
        json_path = json_candidates[0] if len(json_candidates) == 1 else None

    if not csv_path.exists():
        csv_candidates = sorted(directory.glob("*.csv"))
        csv_path = csv_candidates[0] if len(csv_candidates) == 1 else None

    if not excel_path.exists():
        excel_candidates = sorted(
            candidate
            for pattern in ("*.xlsx", "*.xlsm")
            for candidate in directory.glob(pattern)
            if not candidate.name.startswith("~$")
        )
        excel_path = excel_candidates[0] if len(excel_candidates) == 1 else None

    if json_path is None and csv_path is None and excel_path is None:
        raise ValueError("No JSON, CSV, or Excel exports found in the folder.")
    return json_path, csv_path, excel_path


def _existing_fingerprints(transactions: Iterable) -> set:
    return {_fingerprint(tx) for tx in transactions}


def _fingerprint(tx: object) -> Tuple[str, float, str, str, str, str, str, bool]:
    if isinstance(tx, dict):
        datetime_local = str(tx.get("datetime_local", "")).strip()
        amount = float(tx.get("amount", 0))
        tx_type = str(tx.get("type", "")).strip().lower()
        label = str(tx.get("label", "")).strip().lower()
        category = str(tx.get("category", "")).strip().lower()
        notes = str(tx.get("notes", "")).strip().lower()
        recurring_id = str(tx.get("recurring_id", "") or "").strip().lower()
        excluded = _parse_bool(tx.get("excluded_from_averages", False))
    else:
        datetime_local = str(tx.datetime_local)
        amount = float(tx.amount)
        tx_type = str(tx.type).strip().lower()
        label = str(tx.label).strip().lower()
        category = str(tx.category).strip().lower()
        notes = str(tx.notes).strip().lower()
        recurring_id = str(tx.recurring_id or "").strip().lower()
        excluded = bool(getattr(tx, "excluded_from_averages", False))
    return (
        datetime_local,
        round(amount, 2),
        tx_type,
        label,
        category,
        notes,
        recurring_id,
        excluded,
    )


def _load_json(
    path: Path,
) -> Tuple[
    Dict[str, object],
    List[Dict[str, object]],
    List[RecurringCharge],
    List[str],
    List[str],
    int,
    int,
]:
    with path.open("r", encoding="utf-8-sig") as handle:
        payload = json.load(handle)

    if isinstance(payload, list):
        transactions_data = [item for item in payload if isinstance(item, dict)]
        transactions, invalid_transactions = _collect_normalized_rows(
            transactions_data, _normalize_transaction
        )
        messages: List[str] = []
        if invalid_transactions:
            messages.append(
                f"Skipped {invalid_transactions} JSON transaction rows with invalid dates."
            )
        return {}, transactions, [], [], messages, invalid_transactions, 0

    if not isinstance(payload, dict):
        raise ValueError("Invalid JSON format: root must be an object or a list.")

    settings_updates = _map_settings(payload.get("settings", {}))
    transactions_data = payload.get("transactions", []) or []
    if not isinstance(transactions_data, list):
        transactions_data = []
    transactions, invalid_transactions = _collect_normalized_rows(
        [item for item in transactions_data if isinstance(item, dict)],
        _normalize_transaction,
    )
    recurring_data = (
        payload.get("recurring_charges", [])
        or payload.get("recurring", [])
        or []
    )
    if not isinstance(recurring_data, list):
        recurring_data = []
    recurring, invalid_recurring = _collect_normalized_rows(
        [item for item in recurring_data if isinstance(item, dict)],
        _normalize_recurring,
    )
    categories_data = payload.get("categories", []) or []
    if not isinstance(categories_data, list):
        categories_data = []
    categories = [str(item).strip() for item in categories_data if str(item).strip()]
    messages = []
    if invalid_transactions:
        messages.append(
            f"Skipped {invalid_transactions} JSON transaction rows with invalid dates."
        )
    if invalid_recurring:
        messages.append(
            f"Skipped {invalid_recurring} recurring rows with invalid schedule dates."
        )
    return (
        settings_updates,
        transactions,
        recurring,
        categories,
        messages,
        invalid_transactions,
        invalid_recurring,
    )


def _load_csv(path: Path) -> Tuple[List[Dict[str, object]], List[str], int]:
    transactions: List[Dict[str, object]] = []
    invalid_transactions = 0
    with path.open("r", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            try:
                transactions.append(_normalize_transaction(row))
            except ValueError:
                invalid_transactions += 1
    messages: List[str] = []
    if invalid_transactions:
        messages.append(
            f"Skipped {invalid_transactions} CSV rows with invalid dates."
        )
    return transactions, messages, invalid_transactions


def _load_excel(
    path: Path,
) -> Tuple[
    Dict[str, object],
    List[Dict[str, object]],
    List[RecurringCharge],
    List[str],
    List[str],
    int,
    int,
]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel import.")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        export_payload = _load_excel_export_payload(workbook)
        if export_payload is not None:
            (
                settings_updates,
                transactions,
                recurring,
                categories,
                invalid_transactions,
                invalid_recurring,
            ) = export_payload
        else:
            settings_updates = _load_excel_settings(workbook)
            transactions, invalid_transactions = _load_excel_transactions(workbook)
            recurring, invalid_recurring = _load_excel_recurring(workbook)
            categories = _load_excel_categories(workbook)
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()

    if not categories:
        categories = [
            str(tx.get("category", "")).strip()
            for tx in transactions
            if str(tx.get("category", "")).strip()
        ]
        categories.extend(charge.category for charge in recurring if str(charge.category).strip())
    messages: List[str] = []
    if invalid_transactions:
        messages.append(
            f"Skipped {invalid_transactions} Excel rows with invalid dates."
        )
    if invalid_recurring:
        messages.append(
            f"Skipped {invalid_recurring} Excel recurring rows with invalid schedule dates."
        )
    categories = [str(item).strip() for item in categories if str(item).strip()]
    return (
        settings_updates,
        transactions,
        recurring,
        categories,
        messages,
        invalid_transactions,
        invalid_recurring,
    )


def _load_excel_export_payload(
    workbook,
) -> Optional[
    Tuple[
        Dict[str, object],
        List[Dict[str, object]],
        List[RecurringCharge],
        List[str],
        int,
        int,
    ]
]:
    sheet = _find_excel_sheet(workbook, EXCEL_EXPORT_DATA_SHEET_NAMES)
    if sheet is None:
        return None

    raw_sections: Dict[str, List[Tuple[int, str]]] = {}
    for row in sheet.iter_rows(values_only=True):
        if _row_is_blank(row):
            continue
        section = _normalize_lookup_key(_stringify_value(row[0] if len(row) > 0 else ""))
        if not section:
            continue
        if section == "section":
            continue

        second_value = _stringify_value(row[1] if len(row) > 1 else "")
        third_value = _stringify_value(row[2] if len(row) > 2 else "")
        if third_value or _looks_like_excel_chunk_index(second_value):
            raw_json = third_value or second_value
            part_index = _parse_excel_chunk_index(second_value)
        else:
            raw_json = second_value
            part_index = 1
        if not raw_json:
            continue

        raw_sections.setdefault(section, []).append((part_index, raw_json))

    sections: Dict[str, object] = {}
    for section, parts in raw_sections.items():
        raw_json = "".join(chunk for _, chunk in sorted(parts, key=lambda item: item[0]))
        try:
            sections[section] = json.loads(raw_json)
        except json.JSONDecodeError as exc:
            raise ValueError("Excel export data sheet is invalid.") from exc

    settings_data = sections.get("settings", {})
    if not isinstance(settings_data, dict):
        settings_data = {}
    transactions_data = sections.get("transactions", [])
    if not isinstance(transactions_data, list):
        transactions_data = []
    recurring_data = sections.get("recurring charges", [])
    if not isinstance(recurring_data, list):
        recurring_data = []
    categories_data = sections.get("categories", [])
    if not isinstance(categories_data, list):
        categories_data = []

    settings_updates = _map_settings(settings_data)
    transactions, invalid_transactions = _collect_normalized_rows(
        [item for item in transactions_data if isinstance(item, dict)],
        _normalize_transaction,
    )
    recurring, invalid_recurring = _collect_normalized_rows(
        [item for item in recurring_data if isinstance(item, dict)],
        _normalize_recurring,
    )
    categories = [str(item).strip() for item in categories_data if str(item).strip()]
    return (
        settings_updates,
        transactions,
        recurring,
        categories,
        invalid_transactions,
        invalid_recurring,
    )


def _normalize_transaction(raw: Dict[str, object]) -> Dict[str, object]:
    tx_id = _normalize_optional_text(
        raw.get("id") or raw.get("uuid") or raw.get("transaction_id") or raw.get("transactionId")
    )
    datetime_value = (
        raw.get("datetime_local")
        or raw.get("datetime")
        or raw.get("date")
        or raw.get("Date")
    )
    datetime_local = _normalize_datetime(_stringify_value(datetime_value))
    label = _stringify_value(raw.get("label", ""))
    category = _normalize_category_value(_stringify_value(raw.get("category", "")))
    notes = _build_notes(raw, preserve_id=bool(tx_id))
    tx_type = _normalize_transaction_type(raw.get("type", "expense"))
    recurring_id = _normalize_optional_text(
        raw.get("recurring_id") or raw.get("recurringId")
    )
    excluded = _parse_bool(
        raw.get("excluded_from_averages")
        if "excluded_from_averages" in raw
        else raw.get("exclude_from_averages", raw.get("avg_excluded"))
    )
    return {
        "id": tx_id,
        "datetime_local": datetime_local,
        "amount": _safe_float(raw.get("amount", 0), 0.0),
        "type": tx_type,
        "label": label,
        "category": category,
        "notes": notes,
        "recurring_id": recurring_id,
        "excluded_from_averages": excluded,
    }


def _normalize_recurring(raw: Dict[str, object]) -> RecurringCharge:
    charge_id = str(raw.get("id") or "").strip() or str(raw.get("uuid") or "").strip()
    if not charge_id:
        charge_id = f"imported-{datetime.now().timestamp()}-{abs(hash(str(raw))) % 1000000}"

    label = str(raw.get("label", "Recurring")).strip() or "Recurring"
    amount = _safe_float(raw.get("amount", 0), 0.0)
    tx_type = str(raw.get("type", "expense")).strip().lower()
    if tx_type not in {"expense", "income"}:
        tx_type = "expense"
    category = str(raw.get("category", "misc")).strip() or "misc"
    notes = str(raw.get("notes", "")).strip()

    start_date = str(raw.get("start_date") or raw.get("startDate") or "").strip()
    start_date = _normalize_date(start_date, field_name="start_date")
    end_date_raw = str(raw.get("end_date") or raw.get("endDate") or "").strip()
    end_date = (
        _normalize_date(end_date_raw, field_name="end_date")
        if end_date_raw
        else None
    )
    if end_date and end_date < start_date:
        raise ValueError("end_date cannot be earlier than start_date")

    status = str(raw.get("status", "automatic")).strip().lower()
    if status not in {"manual", "automatic", "paused"}:
        status = "automatic"
    frequency = _normalize_frequency(str(raw.get("frequency", "monthly")).strip().lower())
    last_applied = (
        str(raw.get("last_applied") or raw.get("last_applied_date") or raw.get("lastApplied") or "").strip()
        or None
    )

    return RecurringCharge(
        id=charge_id,
        label=label,
        amount=amount,
        type=tx_type,
        category=category,
        notes=notes,
        start_date=start_date,
        end_date=end_date,
        status=status,
        frequency=frequency,
        last_applied=last_applied,
    )


def _recurring_fingerprint(charge: object) -> Tuple[str, float, str, str, str, str, str]:
    if isinstance(charge, dict):
        label = str(charge.get("label", "")).strip().lower()
        amount = float(charge.get("amount", 0) or 0)
        tx_type = str(charge.get("type", "")).strip().lower()
        category = str(charge.get("category", "")).strip().lower()
        start_date = str(charge.get("start_date", "")).strip()
        end_date = str(charge.get("end_date", "") or "").strip()
        frequency = _normalize_frequency(str(charge.get("frequency", "monthly")).strip().lower())
    else:
        label = str(charge.label).strip().lower()
        amount = float(charge.amount)
        tx_type = str(charge.type).strip().lower()
        category = str(charge.category).strip().lower()
        start_date = str(charge.start_date).strip()
        end_date = str(charge.end_date or "").strip()
        frequency = _normalize_frequency(str(charge.frequency).strip().lower())
    return (label, round(amount, 2), tx_type, category, start_date, end_date, frequency)


def _build_notes(raw: Dict[str, object], preserve_id: bool = False) -> str:
    parts = []
    imported_id = raw.get("id")
    if imported_id and not preserve_id:
        parts.append(f"imported_id={imported_id}")
    source = raw.get("source")
    if source:
        parts.append(f"source={source}")
    if "date_estimated" in raw:
        value = raw.get("date_estimated")
        if isinstance(value, str):
            value = value.strip().lower() in {"true", "1", "yes"}
        parts.append(f"date_estimated={str(bool(value)).lower()}")
    notes = str(raw.get("notes", "")).strip()
    meta = "; ".join(parts)
    if notes and meta:
        return f"{notes} | {meta}"
    if notes:
        return notes
    return meta


def _normalize_datetime(value: str) -> str:
    text = (value or "").strip()
    if not text:
        raise ValueError("datetime_local is required")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        try:
            parsed = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
        except ValueError as exc:
            raise ValueError(f"Invalid datetime_local: {value!r}") from exc
    if parsed.tzinfo is None:
        return parsed.replace(microsecond=0).isoformat()
    try:
        converted = parsed.astimezone()
        return converted.replace(tzinfo=None, microsecond=0).isoformat()
    except Exception:
        return parsed.replace(tzinfo=None, microsecond=0).isoformat()


def _normalize_date(value: str, field_name: str = "date") -> str:
    text = (value or "").strip()
    if not text:
        raise ValueError(f"{field_name} is required")
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"Invalid {field_name}: {value!r}") from exc


def _normalize_optional_text(value: object) -> Optional[str]:
    text = str(value).strip() if value is not None else ""
    return text or None


def _collect_normalized_rows(raw_rows, normalizer):
    normalized = []
    invalid_count = 0
    for row in raw_rows:
        try:
            normalized.append(normalizer(row))
        except ValueError:
            invalid_count += 1
    return normalized, invalid_count


def _load_excel_transactions(workbook) -> Tuple[List[Dict[str, object]], int]:
    sheet = _find_excel_sheet(workbook, EXCEL_TRANSACTION_SHEET_NAMES)
    if sheet is None:
        return [], 0

    header_map = None
    transactions: List[Dict[str, object]] = []
    invalid_count = 0
    for row in sheet.iter_rows(values_only=True):
        if _row_is_blank(row):
            continue
        if header_map is None:
            header_map = _map_excel_headers(
                row,
                EXCEL_TRANSACTION_HEADER_MAP,
                required_fields={"datetime_local", "amount"},
            )
            if header_map:
                continue
            raise ValueError("Excel Transactions sheet has an unsupported header row.")
        raw = _excel_row_to_dict(row, header_map)
        if not raw:
            continue
        try:
            transactions.append(_normalize_transaction(raw))
        except ValueError:
            invalid_count += 1
    return transactions, invalid_count


def _load_excel_settings(workbook) -> Dict[str, object]:
    sheet = _find_excel_sheet(workbook, EXCEL_SETTINGS_SHEET_NAMES)
    if sheet is None:
        return {}

    settings_updates: Dict[str, object] = {}
    for row in sheet.iter_rows(values_only=True):
        if _row_is_blank(row):
            continue
        label = _normalize_lookup_key(_stringify_value(row[0] if len(row) > 0 else ""))
        value = row[1] if len(row) > 1 else None
        if label in {"field", "champ", "value", "valeur"}:
            continue
        setting_key = EXCEL_SETTINGS_LABEL_MAP.get(label)
        if not setting_key:
            continue
        normalized = _normalize_excel_setting(setting_key, value)
        if normalized is not None:
            settings_updates[setting_key] = normalized
    return settings_updates


def _load_excel_recurring(workbook) -> Tuple[List[RecurringCharge], int]:
    sheet = _find_excel_sheet(workbook, EXCEL_RECURRING_SHEET_NAMES)
    if sheet is None:
        return [], 0

    header_map = None
    recurring: List[RecurringCharge] = []
    invalid_count = 0
    for row in sheet.iter_rows(values_only=True):
        if _row_is_blank(row):
            continue
        if header_map is None:
            header_map = _map_excel_headers(
                row,
                EXCEL_RECURRING_HEADER_MAP,
                required_fields={"amount", "start_date"},
            )
            if header_map:
                continue
            raise ValueError("Excel recurring sheet has an unsupported header row.")
        raw = _excel_row_to_dict(row, header_map)
        if not raw:
            continue
        try:
            recurring.append(_normalize_recurring(raw))
        except ValueError:
            invalid_count += 1
    return recurring, invalid_count


def _load_excel_categories(workbook) -> List[str]:
    sheet = _find_excel_sheet(workbook, EXCEL_CATEGORY_SHEET_NAMES)
    if sheet is None:
        return []

    categories: List[str] = []
    header_skipped = False
    for row in sheet.iter_rows(values_only=True):
        if _row_is_blank(row):
            continue
        value = _normalize_category_value(_stringify_value(row[0] if len(row) > 0 else ""))
        if not value:
            continue
        if not header_skipped and _normalize_lookup_key(value) in {"category", "categorie"}:
            header_skipped = True
            continue
        categories.append(value)
        header_skipped = True
    return categories


def _find_excel_sheet(workbook, candidate_names: set):
    for sheet_name in workbook.sheetnames:
        if _normalize_lookup_key(sheet_name) in candidate_names:
            return workbook[sheet_name]
    return None


def _map_excel_headers(
    row: Tuple[object, ...],
    aliases: Dict[str, str],
    required_fields: Optional[set] = None,
) -> Optional[Dict[int, str]]:
    header_map: Dict[int, str] = {}
    for index, cell in enumerate(row):
        normalized = _normalize_lookup_key(_stringify_value(cell))
        if not normalized:
            continue
        target = aliases.get(normalized)
        if target:
            header_map[index] = target
    required = required_fields or {"datetime_local", "amount"}
    if not required.issubset(set(header_map.values())):
        return None
    return header_map


def _excel_row_to_dict(
    row: Tuple[object, ...],
    header_map: Dict[int, str],
) -> Dict[str, object]:
    raw: Dict[str, object] = {}
    for index, field_name in header_map.items():
        if index >= len(row):
            continue
        value = row[index]
        if value is None:
            continue
        if field_name == "datetime_local":
            raw[field_name] = _excel_datetime_text(value)
        else:
            raw[field_name] = value
    return raw


def _normalize_excel_setting(setting_key: str, value: object) -> Optional[object]:
    if value is None:
        return None
    if setting_key in EXCEL_INT_SETTINGS:
        return int(round(_safe_float(value, 0.0)))
    if setting_key in EXCEL_FLOAT_SETTINGS:
        return _safe_float(value, 0.0)
    if setting_key == "language":
        text = _stringify_value(value).upper()
        return text if text in {"EN", "FR"} else None
    return _stringify_value(value)


def _looks_like_excel_chunk_index(value: str) -> bool:
    if not value:
        return False
    return bool(re.fullmatch(r"\d+", value))


def _parse_excel_chunk_index(value: str) -> int:
    if not _looks_like_excel_chunk_index(value):
        return 1
    try:
        return max(1, int(value))
    except ValueError:
        return 1


def _normalize_transaction_type(value: object) -> str:
    text = _normalize_lookup_key(_stringify_value(value or "expense"))
    return TRANSACTION_TYPE_ALIASES.get(text, "expense")


def _normalize_category_value(value: object) -> str:
    text = _stringify_value(value)
    if not text:
        return ""
    return CATEGORY_LOOKUP.get(_normalize_lookup_key(text), text)


def _normalize_lookup_key(value: str) -> str:
    return " ".join(value.replace("_", " ").strip().lower().split())


def _stringify_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    return str(value).strip()


def _excel_datetime_text(value: object) -> str:
    return _stringify_value(value)


def _row_is_blank(row: Tuple[object, ...]) -> bool:
    return not any(_stringify_value(cell) for cell in row)


def _safe_float(value: object, default: float = 0.0) -> float:
    text = str(value).strip() if value is not None else ""
    if text:
        # Accept both "1,234.56" and "1234,56" styles from exported spreadsheets.
        if "," in text and "." in text:
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    try:
        parsed = float(text or 0)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(parsed):
        return default
    return parsed


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "on"}


def _normalize_frequency(value: str) -> str:
    if value in {"daily", "weekly", "monthly"}:
        return value
    return "monthly"


def _map_settings(raw: Dict[str, object]) -> Dict[str, object]:
    updates: Dict[str, object] = {}
    mapping = {
        "starting_balance": "starting_balance",
        "refunds_total": "refunds_total",
        "refunds_received_total": "refunds_total",
        "counted_balance": "counted_balance",
        "counted_balance_latest": "counted_balance",
        "plan_months_total": "plan_months_total",
        "months_elapsed": "months_elapsed",
        "rent_base_monthly": "rent_base_monthly",
        "rent_monthly_manual": "rent_monthly_manual",
        "food_house_monthly": "food_house_monthly",
        "misc_monthly": "misc_monthly",
        "medical_monthly": "medical_monthly",
        "school_monthly": "school_monthly",
        "household_monthly": "household_monthly",
        "health_monthly": "health_monthly",
        "auto_misc_enabled": "auto_misc_enabled",
        "auto_misc_categories": "auto_misc_categories",
        "auto_include_recurring": "auto_include_recurring",
        "auto_include_current_month": "auto_include_current_month",
        "auto_window_months": "auto_window_months",
        "auto_weighted": "auto_weighted",
        "auto_weight_half_life_months": "auto_weight_half_life_months",
        "extra_monthly": "extra_monthly",
        "rent_surcharge_amount": "rent_surcharge_amount",
        "rent_surcharge_months": "rent_surcharge_months",
        "campus_reference_total": "campus_reference_total",
        "language": "language",
        "theme": "theme",
        "font_family": "font_family",
        "accent_color": "accent_color",
    }
    for key, target in mapping.items():
        if key in raw and raw[key] is not None:
            updates[target] = raw[key]
    return updates
