from __future__ import annotations

import io
import json
from dataclasses import asdict
from datetime import datetime
from typing import List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from xml.sax.saxutils import escape
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover - optional for Excel export
    Workbook = None
    Font = None

from . import charts
from .calculations import group_transactions_by_month
from .db import RecurringCharge, Settings, TZ_NAME, Transaction
from .localization import translated_category


LABELS = {
    "EN": {
        "title": "Offline Budget Report",
        "inputs": "Inputs and Settings",
        "overview": "Overview",
        "highlights": "Key highlights",
        "balances_section": "Balances",
        "plan_section": "Plan & forecast",
        "coverage_section": "Coverage",
        "notes_section": "How to read this report",
        "calculations": "Calculations",
        "transactions": "Monthly Transactions",
        "totals": "Totals and Percentages",
        "graphs": "Graphs",
        "date": "Generated",
        "starting_balance": "Starting balance",
        "refunds_total": "Refunds total",
        "plan_months_total": "Plan months total",
        "months_elapsed": "Months elapsed",
        "rent_base_monthly": "Rent (managed via recurring)",
        "rent_monthly_manual": "Rent monthly (manual)",
        "food_house_monthly": "Food house monthly",
        "medical_monthly": "Medical monthly",
        "school_monthly": "School monthly",
        "household_monthly": "Household monthly",
        "health_monthly": "Health monthly",
        "misc_monthly": "Misc monthly",
        "extra_monthly": "Extra monthly (12 mo)",
        "campus_reference_total": "Campus reference total",
        "estimated_balance": "Estimated balance",
        "total_expenses": "Total expenses",
        "total_income": "Total income",
        "planned_total": "Planned total (manual + recurring)",
        "predicted_total": "Predicted total (auto avg)",
        "planned_remaining_total": "Planned remaining total",
        "predicted_remaining_total": "Predicted remaining total",
        "essential_budget_total_planned": "Essential budget total (planned)",
        "essential_budget_total_predicted": "Essential budget total (predicted)",
        "base_rent_total": "Rent plan total (projected)",
        "food_total": "Food total (projected)",
        "essential_month_used_pct_planned": "Essential used % (month, planned)",
        "essential_month_used_pct_predicted": "Essential used % (month, predicted)",
        "essential_year_used_pct_planned": "Essential used % (year, planned)",
        "essential_year_used_pct_predicted": "Essential used % (year, predicted)",
        "rent_month_used_pct_planned": "Rent used % (month, planned)",
        "rent_month_used_pct_predicted": "Rent used % (month, predicted)",
        "rent_year_used_pct_planned": "Rent used % (year, planned)",
        "rent_year_used_pct_predicted": "Rent used % (year, predicted)",
        "savings_vs_campus_planned": "Savings vs campus (planned)",
        "savings_vs_campus_predicted": "Savings vs campus (predicted)",
        "rent_paid_to_date": "Rent paid to date",
        "months_remaining": "Months remaining",
        "rent_remaining_total": "Recurring remaining total",
        "cap_nonrent_per_month": "Cap non-recurring per month",
        "coverage_display": "Coverage (school year, projected)",
        "coverage_12mo_display": "Coverage (12 months, projected)",
        "language": "Language",
        "sheet_summary": "Summary",
        "sheet_settings": "Settings",
        "sheet_transactions": "Transactions",
        "sheet_monthly": "MonthlyStats",
        "metric": "Metric",
        "value": "Value",
        "field": "Field",
        "transactions_header": ["Date", "Type", "Amount", "Label", "Category", "Notes"],
        "monthly_header": ["Month", "Total expenses", "Delta $", "Delta %"],
        "settings_recurring_label": "rent (recurring)",
        "settings_recurring_value": "managed in recurring charges",
        "overview_blurb": "This report summarizes your balances, monthly spending, and plan outlook.",
        "notes_blurb": "Planned values use your manual settings plus recurring charges. Predicted values use recent averages.",
    },
    "FR": {
        "title": "Rapport Budget Hors Ligne",
        "inputs": "Entrees et Reglages",
        "overview": "Apercu",
        "highlights": "Faits saillants",
        "balances_section": "Soldes",
        "plan_section": "Plan et previsions",
        "coverage_section": "Couverture",
        "notes_section": "Comment lire ce rapport",
        "calculations": "Calculs",
        "transactions": "Transactions Mensuelles",
        "totals": "Totaux et Pourcentages",
        "graphs": "Graphiques",
        "date": "Genere",
        "starting_balance": "Solde initial",
        "refunds_total": "Total remboursements",
        "plan_months_total": "Total mois plan",
        "months_elapsed": "Mois ecoules",
        "rent_base_monthly": "Loyer (charges recurrentes)",
        "rent_monthly_manual": "Loyer mensuel (manuel)",
        "food_house_monthly": "Nourriture maison mensuel",
        "medical_monthly": "Medical mensuel",
        "school_monthly": "Ecole mensuel",
        "household_monthly": "Maison mensuel",
        "health_monthly": "Sante mensuel",
        "misc_monthly": "Divers mensuel",
        "extra_monthly": "Extra mensuel (12 mois)",
        "campus_reference_total": "Reference campus",
        "estimated_balance": "Solde estime",
        "total_expenses": "Total depenses",
        "total_income": "Total revenus",
        "planned_total": "Total planifie (manuel + recurrent)",
        "predicted_total": "Total predit (moyenne auto)",
        "planned_remaining_total": "Total planifie restant",
        "predicted_remaining_total": "Total predit restant",
        "essential_budget_total_planned": "Budget essentiel total (planifie)",
        "essential_budget_total_predicted": "Budget essentiel total (predite)",
        "base_rent_total": "Total loyer plan (projete)",
        "food_total": "Total nourriture (projete)",
        "essential_month_used_pct_planned": "% essentiel utilise (mois, planifie)",
        "essential_month_used_pct_predicted": "% essentiel utilise (mois, predit)",
        "essential_year_used_pct_planned": "% essentiel utilise (annee, planifie)",
        "essential_year_used_pct_predicted": "% essentiel utilise (annee, predit)",
        "rent_month_used_pct_planned": "% loyer utilise (mois, planifie)",
        "rent_month_used_pct_predicted": "% loyer utilise (mois, predit)",
        "rent_year_used_pct_planned": "% loyer utilise (annee, planifie)",
        "rent_year_used_pct_predicted": "% loyer utilise (annee, predit)",
        "savings_vs_campus_planned": "Economie vs campus (planifiee)",
        "savings_vs_campus_predicted": "Economie vs campus (predite)",
        "rent_paid_to_date": "Loyer paye a ce jour",
        "months_remaining": "Mois restants",
        "rent_remaining_total": "Total charges recurrentes restantes",
        "cap_nonrent_per_month": "Cap hors recurrent par mois",
        "coverage_display": "Couverture (annee scolaire, projete)",
        "coverage_12mo_display": "Couverture (12 mois, projete)",
        "language": "Langue",
        "sheet_summary": "Resume",
        "sheet_settings": "Reglages",
        "sheet_transactions": "Transactions",
        "sheet_monthly": "StatistiquesMensuelles",
        "metric": "Mesure",
        "value": "Valeur",
        "field": "Champ",
        "transactions_header": ["Date", "Type", "Montant", "Libelle", "Categorie", "Notes"],
        "monthly_header": ["Mois", "Total depenses", "Delta $", "Delta %"],
        "settings_recurring_label": "loyer (recurrent)",
        "settings_recurring_value": "gere via charges recurrentes",
        "overview_blurb": "Ce rapport resume vos soldes, depenses mensuelles et projections.",
        "notes_blurb": "Les valeurs planifiees utilisent vos reglages manuels et charges recurrentes. Les valeurs predites utilisent les moyennes recentes.",
    },
}

PAGE_MARGIN = 36
CONTENT_WIDTH = letter[0] - (PAGE_MARGIN * 2)
EXCEL_EXPORT_DATA_SHEET = "_export_data"
EXCEL_EXPORT_MAX_CELL_CHARS = 30000
VALUE_LABELS = {
    "expense": {"EN": "Expense", "FR": "Depense"},
    "income": {"EN": "Income", "FR": "Revenu"},
}


def _type_label(value: str, language: str) -> str:
    key = str(value or "").strip().lower()
    text = VALUE_LABELS.get(key)
    if not text:
        return str(value)
    lang = (language or "EN").upper()
    return text.get(lang, text["EN"])


def export_pdf(
    output_path: str,
    settings: Settings,
    transactions: List[Transaction],
    summary: dict,
    monthly_stats: List[dict],
) -> None:
    labels = LABELS.get(settings.language, LABELS["EN"])
    styles = getSampleStyleSheet()
    table_text = styles["BodyText"].clone("TableText")
    table_text.fontSize = 8
    table_text.leading = 10
    story = []

    story.append(Paragraph(labels["title"], styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            f"{labels['date']}: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))
    story.append(Paragraph(labels["overview"], styles["Heading1"]))
    story.append(Paragraph(labels.get("overview_blurb", ""), styles["BodyText"]))
    story.append(Spacer(1, 8))

    highlight_rows = [
        [labels["estimated_balance"], f"{summary['estimated_balance']:.2f}"],
        [labels["total_expenses"], f"{summary['total_expenses']:.2f}"],
        [labels["planned_total"], f"{summary['plan_budget_total']:.2f}"],
        [labels["predicted_total"], f"{summary['projected_final_expenses']:.2f}"],
        [labels["coverage_display"], summary.get("coverage_display") or ""],
        [labels["coverage_12mo_display"], summary.get("coverage_12mo_display") or ""],
    ]
    story.append(Paragraph(labels["highlights"], styles["Heading2"]))
    story.append(_styled_table(highlight_rows, col_widths=_two_col_widths()))
    story.append(Spacer(1, 8))
    story.append(Paragraph(labels.get("notes_section", "How to read this report"), styles["Heading2"]))
    story.append(Paragraph(labels.get("notes_blurb", ""), styles["BodyText"]))
    story.append(PageBreak())

    story.append(Paragraph(labels["inputs"], styles["Heading1"]))
    settings_rows = [
        [labels["starting_balance"], f"{settings.starting_balance:.2f}"],
        [labels["refunds_total"], f"{settings.refunds_total:.2f}"],
        [labels["plan_months_total"], str(settings.plan_months_total)],
        [labels["months_elapsed"], str(settings.months_elapsed)],
        [labels.get("rent_monthly_manual", "Rent monthly (manual)"), f"{settings.rent_monthly_manual:.2f}"],
        [labels["food_house_monthly"], f"{settings.food_house_monthly:.2f}"],
        [labels.get("medical_monthly", "Medical monthly"), f"{settings.medical_monthly:.2f}"],
        [labels.get("school_monthly", "School monthly"), f"{settings.school_monthly:.2f}"],
        [labels.get("household_monthly", "Household monthly"), f"{settings.household_monthly:.2f}"],
        [labels.get("health_monthly", "Health monthly"), f"{settings.health_monthly:.2f}"],
        [labels["misc_monthly"], f"{settings.misc_monthly:.2f}"],
        [labels["extra_monthly"], f"{settings.extra_monthly:.2f}"],
        [labels["campus_reference_total"], f"{settings.campus_reference_total:.2f}"],
        [labels["rent_base_monthly"], "-"],
    ]
    story.append(
        _styled_table(
            settings_rows,
            col_widths=_two_col_widths(),
        )
    )
    story.append(PageBreak())

    story.append(Paragraph(labels["calculations"], styles["Heading1"]))
    balances_rows = [
        [labels["estimated_balance"], f"{summary['estimated_balance']:.2f}"],
        [labels["total_expenses"], f"{summary['total_expenses']:.2f}"],
        [labels["total_income"], f"{summary['total_income']:.2f}"],
    ]
    plan_rows = [
        [labels["planned_total"], f"{summary['plan_budget_total']:.2f}"],
        [labels["predicted_total"], f"{summary['projected_final_expenses']:.2f}"],
        [labels["planned_remaining_total"], f"{summary['planned_remaining_total']:.2f}"],
        [labels["predicted_remaining_total"], f"{summary['predicted_remaining_total']:.2f}"],
        [labels["essential_budget_total_planned"], f"{summary.get('essential_budget_total_planned', 0.0):.2f}"],
        [labels["essential_budget_total_predicted"], f"{summary.get('essential_budget_total_predicted', 0.0):.2f}"],
        [labels["savings_vs_campus_planned"], f"{summary.get('savings_vs_campus_planned', 0.0):.2f}"],
        [labels["savings_vs_campus_predicted"], f"{summary.get('savings_vs_campus_predicted', 0.0):.2f}"],
    ]
    coverage_rows = [
        [labels["months_remaining"], str(summary["months_remaining"])],
        [labels["rent_remaining_total"], f"{summary['rent_remaining_total']:.2f}"],
        [
            labels["cap_nonrent_per_month"],
            "" if summary["cap_nonrent_per_month"] is None else f"{summary['cap_nonrent_per_month']:.2f}",
        ],
        [labels["coverage_display"], summary["coverage_display"] or ""],
        [labels["coverage_12mo_display"], summary.get("coverage_12mo_display") or ""],
    ]
    story.append(Paragraph(labels["balances_section"], styles["Heading2"]))
    story.append(_styled_table(balances_rows, col_widths=_two_col_widths()))
    story.append(Spacer(1, 8))
    story.append(Paragraph(labels["plan_section"], styles["Heading2"]))
    story.append(_styled_table(plan_rows, col_widths=_two_col_widths()))
    story.append(Spacer(1, 8))
    story.append(Paragraph(labels["coverage_section"], styles["Heading2"]))
    story.append(_styled_table(coverage_rows, col_widths=_two_col_widths()))
    story.append(PageBreak())

    story.append(Paragraph(labels["transactions"], styles["Heading1"]))
    grouped = group_transactions_by_month(transactions)
    for idx, month_key in enumerate(sorted(grouped.keys())):
        if idx > 0:
            story.append(PageBreak())
        story.append(Paragraph(month_key, styles["Heading2"]))
        tx_rows = _build_transaction_rows(
            grouped[month_key],
            table_text,
            headers=labels.get("transactions_header"),
            language=settings.language,
        )
        story.append(
            _styled_table(
                tx_rows,
                header=True,
                col_widths=_transaction_col_widths(),
                repeat_rows=1,
                extra_styles=[
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ],
            )
        )
    story.append(PageBreak())

    story.append(Paragraph(labels["totals"], styles["Heading1"]))
    totals_rows = [
        [labels["total_expenses"], f"{summary['total_expenses']:.2f}"],
        [labels["total_income"], f"{summary['total_income']:.2f}"],
        [
            labels["essential_month_used_pct_planned"],
            f"{summary.get('essential_month_used_pct_planned', 0.0):.2f}%",
        ],
        [
            labels["essential_month_used_pct_predicted"],
            f"{summary.get('essential_month_used_pct_predicted', 0.0):.2f}%",
        ],
        [
            labels["essential_year_used_pct_planned"],
            f"{summary.get('essential_year_used_pct_planned', 0.0):.2f}%",
        ],
        [
            labels["essential_year_used_pct_predicted"],
            f"{summary.get('essential_year_used_pct_predicted', 0.0):.2f}%",
        ],
        [
            labels["rent_month_used_pct_planned"],
            f"{summary.get('rent_month_used_pct_planned', 0.0):.2f}%",
        ],
        [
            labels["rent_month_used_pct_predicted"],
            f"{summary.get('rent_month_used_pct_predicted', 0.0):.2f}%",
        ],
        [
            labels["rent_year_used_pct_planned"],
            f"{summary.get('rent_year_used_pct_planned', 0.0):.2f}%",
        ],
        [
            labels["rent_year_used_pct_predicted"],
            f"{summary.get('rent_year_used_pct_predicted', 0.0):.2f}%",
        ],
        [labels["rent_paid_to_date"], f"{summary['rent_paid_to_date']:.2f}"],
    ]
    story.append(
        _styled_table(
            totals_rows,
            col_widths=_two_col_widths(),
        )
    )
    story.append(PageBreak())

    story.append(Paragraph(labels["graphs"], styles["Heading1"]))
    story.extend(_chart_flowables(transactions, settings, summary, monthly_stats))

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=PAGE_MARGIN,
        rightMargin=PAGE_MARGIN,
        topMargin=PAGE_MARGIN,
        bottomMargin=PAGE_MARGIN,
    )
    doc.build(story)


def export_excel(
    output_path: str,
    settings: Settings,
    transactions: List[Transaction],
    summary: dict,
    monthly_stats: List[dict],
    recurring_charges: Optional[List[RecurringCharge]] = None,
    categories: Optional[List[str]] = None,
) -> None:
    if Workbook is None or Font is None:
        raise RuntimeError("openpyxl is required for Excel export.")
    labels = LABELS.get(settings.language, LABELS["EN"])
    bold = Font(bold=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = labels.get("sheet_summary", "Summary")

    summary_rows = [
        [labels.get("metric", "Metric"), labels.get("value", "Value")],
        [labels["estimated_balance"], summary["estimated_balance"]],
        [labels["total_expenses"], summary["total_expenses"]],
        [labels["total_income"], summary["total_income"]],
        [labels["planned_total"], summary["plan_budget_total"]],
        [labels["predicted_total"], summary["projected_final_expenses"]],
        [labels["planned_remaining_total"], summary["planned_remaining_total"]],
        [labels["predicted_remaining_total"], summary["predicted_remaining_total"]],
        [labels["base_rent_total"], summary["base_rent_total"]],
        [labels["food_total"], summary["food_total"]],
        [labels["essential_budget_total_planned"], summary.get("essential_budget_total_planned")],
        [labels["essential_budget_total_predicted"], summary.get("essential_budget_total_predicted")],
        [labels["savings_vs_campus_planned"], summary.get("savings_vs_campus_planned")],
        [labels["savings_vs_campus_predicted"], summary.get("savings_vs_campus_predicted")],
        [labels["rent_paid_to_date"], summary["rent_paid_to_date"]],
        [labels["essential_month_used_pct_planned"], summary.get("essential_month_used_pct_planned")],
        [labels["essential_month_used_pct_predicted"], summary.get("essential_month_used_pct_predicted")],
        [labels["essential_year_used_pct_planned"], summary.get("essential_year_used_pct_planned")],
        [labels["essential_year_used_pct_predicted"], summary.get("essential_year_used_pct_predicted")],
        [labels["rent_month_used_pct_planned"], summary.get("rent_month_used_pct_planned")],
        [labels["rent_month_used_pct_predicted"], summary.get("rent_month_used_pct_predicted")],
        [labels["rent_year_used_pct_planned"], summary.get("rent_year_used_pct_planned")],
        [labels["rent_year_used_pct_predicted"], summary.get("rent_year_used_pct_predicted")],
        [labels["months_remaining"], summary["months_remaining"]],
        [labels["rent_remaining_total"], summary["rent_remaining_total"]],
        [labels["cap_nonrent_per_month"], summary["cap_nonrent_per_month"]],
        [labels["coverage_display"], summary["coverage_display"]],
        [labels["coverage_12mo_display"], summary.get("coverage_12mo_display")],
    ]
    _write_sheet(summary_sheet, summary_rows, bold)

    settings_sheet = workbook.create_sheet(labels.get("sheet_settings", "Settings"))
    settings_rows = [
        [labels.get("field", "Field"), labels.get("value", "Value")],
        [labels["starting_balance"], settings.starting_balance],
        [labels["refunds_total"], settings.refunds_total],
        [labels["plan_months_total"], settings.plan_months_total],
        [labels["months_elapsed"], settings.months_elapsed],
        [labels.get("rent_monthly_manual", "Rent monthly (manual)"), settings.rent_monthly_manual],
        [labels["food_house_monthly"], settings.food_house_monthly],
        [labels.get("medical_monthly", "Medical monthly"), settings.medical_monthly],
        [labels.get("school_monthly", "School monthly"), settings.school_monthly],
        [labels.get("household_monthly", "Household monthly"), settings.household_monthly],
        [labels.get("health_monthly", "Health monthly"), settings.health_monthly],
        [labels["misc_monthly"], settings.misc_monthly],
        [labels["extra_monthly"], settings.extra_monthly],
        [labels["campus_reference_total"], settings.campus_reference_total],
        [labels.get("settings_recurring_label", "rent (recurring)"),
         labels.get("settings_recurring_value", "managed in recurring charges")],
        [labels["language"], settings.language],
    ]
    _write_sheet(settings_sheet, settings_rows, bold)

    tx_sheet = workbook.create_sheet(labels.get("sheet_transactions", "Transactions"))
    tx_rows = [labels.get("transactions_header", ["Date", "Type", "Amount", "Label", "Category", "Notes"])]
    for tx in transactions:
        tx_rows.append(
            [
                tx.datetime_local,
                _type_label(tx.type, settings.language),
                tx.amount,
                tx.label,
                translated_category(tx.category, settings.language),
                tx.notes,
            ]
        )
    _write_sheet(tx_sheet, tx_rows, bold)

    monthly_sheet = workbook.create_sheet(labels.get("sheet_monthly", "MonthlyStats"))
    monthly_rows = [labels.get("monthly_header", ["Month", "Total expenses", "Delta $", "Delta %"])]
    for item in monthly_stats:
        monthly_rows.append(
            [
                item["month"],
                item["total_expenses"],
                item["delta_amount"],
                item["delta_pct"],
            ]
        )
    _write_sheet(monthly_sheet, monthly_rows, bold)

    export_data_sheet = workbook.create_sheet(EXCEL_EXPORT_DATA_SHEET)
    export_data_sheet.sheet_state = "hidden"
    export_sections = [
        ("settings", json.dumps(asdict(settings), ensure_ascii=False)),
        ("transactions", json.dumps([asdict(tx) for tx in transactions], ensure_ascii=False)),
        (
            "recurring_charges",
            json.dumps([asdict(charge) for charge in (recurring_charges or [])], ensure_ascii=False),
        ),
        (
            "categories",
            json.dumps(
                sorted(
                    {str(name).strip() for name in (categories or []) if str(name).strip()},
                    key=lambda value: value.lower(),
                ),
                ensure_ascii=False,
            ),
        ),
    ]
    export_payload_rows = [["section", "part", "json"]]
    for section_name, raw_json in export_sections:
        for part_index, chunk in enumerate(_chunk_excel_export_json(raw_json), start=1):
            export_payload_rows.append([section_name, part_index, chunk])
    _write_sheet(export_data_sheet, export_payload_rows, bold)

    workbook.save(output_path)


def export_json(
    output_path: str,
    settings: Settings,
    transactions: List[Transaction],
    recurring_charges: List[RecurringCharge],
    categories: List[str],
) -> None:
    payload = {
        "app": "OfflineBudgetTracker",
        "exported_at": datetime.now().replace(microsecond=0).isoformat(),
        "timezone": TZ_NAME,
        "settings": asdict(settings),
        "transactions": [asdict(tx) for tx in transactions],
        "recurring_charges": [asdict(charge) for charge in recurring_charges],
        "categories": sorted(
            {str(name).strip() for name in categories if str(name).strip()},
            key=lambda value: value.lower(),
        ),
    }
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)


def _styled_table(
    rows,
    header: bool = False,
    col_widths: Optional[List[float]] = None,
    repeat_rows: int = 0,
    extra_styles: Optional[List[tuple]] = None,
) -> Table:
    table = Table(
        rows,
        hAlign="LEFT",
        colWidths=col_widths,
        repeatRows=repeat_rows,
        splitByRow=1,
    )
    style_cmds = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#666666")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]
    if header:
        style_cmds.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d0d7de")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
            ]
        )
    else:
        style_cmds.append(("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ffffff")))
    if extra_styles:
        style_cmds.extend(extra_styles)
    table.setStyle(TableStyle(style_cmds))
    return table


def _two_col_widths() -> List[float]:
    return [CONTENT_WIDTH * 0.62, CONTENT_WIDTH * 0.38]


def _transaction_col_widths() -> List[float]:
    weights = [0.22, 0.1, 0.12, 0.18, 0.12, 0.26]
    return [CONTENT_WIDTH * weight for weight in weights]


def _safe_paragraph(text: str, style) -> Paragraph:
    cleaned = escape(text or "").replace("\n", "<br/>")
    return Paragraph(cleaned, style)


def _build_transaction_rows(
    transactions: List[Transaction],
    style,
    headers: Optional[List[str]] = None,
    language: str = "EN",
) -> List[List[object]]:
    header_row = headers or ["Date", "Type", "Amount", "Label", "Category", "Notes"]
    rows: List[List[object]] = [header_row]
    for tx in transactions:
        rows.append(
            [
                str(tx.datetime_local),
                _type_label(str(tx.type), language),
                f"{tx.amount:.2f}",
                _safe_paragraph(tx.label, style),
                _safe_paragraph(translated_category(tx.category, language), style),
                _safe_paragraph(tx.notes, style),
            ]
        )
    return rows


def _chart_flowables(
    transactions: List[Transaction],
    settings: Settings,
    summary: dict,
    monthly_stats: List[dict],
):
    flowables = []
    charts_data = [
        charts.render_monthly_spending(transactions, settings.language),
        charts.render_cumulative_vs_plan(monthly_stats, settings.language),
        charts.render_running_balance(transactions, settings, settings.language),
        charts.render_category_averages(transactions, settings.language),
    ]
    for img_bytes in charts_data:
        img = Image(io.BytesIO(img_bytes), width=480, height=240)
        flowables.append(img)
        flowables.append(Spacer(1, 12))
    return flowables


def _chunk_excel_export_json(raw_json: str) -> List[str]:
    if not raw_json:
        return [""]
    return [
        raw_json[index : index + EXCEL_EXPORT_MAX_CELL_CHARS]
        for index in range(0, len(raw_json), EXCEL_EXPORT_MAX_CELL_CHARS)
    ]


def _write_sheet(sheet, rows, bold_font: Font) -> None:
    for row_index, row in enumerate(rows, start=1):
        cleaned = ["" if value is None else value for value in row]
        sheet.append(cleaned)
        if row_index == 1:
            for col_index in range(1, len(row) + 1):
                sheet.cell(row=row_index, column=col_index).font = bold_font
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
